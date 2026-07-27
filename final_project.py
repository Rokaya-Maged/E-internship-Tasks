import paramiko
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def read_routers(file_path):
    routers = []

    wb = load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, ip = row
        routers.append({
            "Name": name,
            "IP": ip
        })

    return routers


def collect_data(file_path, new_version):
    routers = read_routers(file_path)

    software_pattern = re.compile(r"Software\s+\(([^)]+)\)")
    version_pattern = re.compile(r"Version\s+([\S]+)")
    release_pattern = re.compile(r"RELEASE SOFTWARE\s+\(([^)]+)\)")

    results = []

    for router in routers:
        paramiko.Transport._preferred_kex = ('diffie-hellman-group1-sha1',)
        paramiko.Transport._preferred_keys = ('ssh-rsa',)
        paramiko.Transport._preferred_ciphers = ('aes128-cbc', '3des-cbc')
        paramiko.Transport._preferred_macs = ('hmac-sha1', 'hmac-md5')

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        ssh.connect(
            hostname=router["IP"],
            username="admin",
            password="admin123",
        )

        stdin, stdout, stderr = ssh.exec_command("show version")
        output = stdout.read().decode()

        first_line = output.splitlines()[0]

        software_match = re.search(r"Software\s+\(([^)]+)\)", first_line)
        version_match = re.search(r"Version\s+([\d.]+)", first_line)
        release_match = re.search(r"RELEASE SOFTWARE\s+\(([^)]+)\)", first_line)

        software = software_match.group(1) if software_match else "N/A"
        version = version_match.group(1) if version_match else "N/A"
        release = release_match.group(1) if release_match else "N/A"

        version_need = "No" if version == new_version else "Yes"

        results.append({
            "Name": router["Name"],
            "IP": router["IP"],
            "Software": software,
            "Release": release,
            "Version": version,
            "Version Need": version_need
        })

        ssh.close()

    return results


def export_to_excel(results, output_file):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results_task3"

    headers = ["Name", "IP", "Software", "Release", "Version", "Version Need"]
    ws.append(headers)

    for row_index, item in enumerate(results, start=2):
        ws.append([item[h] for h in headers])

        cell = ws.cell(row=row_index, column=6)  # Column 6 = Version Need

        if item["Version Need"] == "No":
            cell.fill = green_fill
        else:
            cell.fill = red_fill

    wb.save(output_file)